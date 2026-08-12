import os
import uuid
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

class ExcelService:
    def __init__(self):
        base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        self.data_dir = os.path.join(base_dir, "data")
        os.makedirs(self.data_dir, exist_ok=True)
        
        # Design tokens
        self.primary_color = "2B3A42" # Dark Blue/Gray
        self.accent_color = "3F5765" # Lighter Blue/Gray
        self.header_fill = PatternFill(start_color=self.primary_color, end_color=self.primary_color, fill_type="solid")
        self.total_fill = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")
        
        self.header_font = Font(color="FFFFFF", bold=True, size=12)
        self.title_font = Font(color=self.primary_color, bold=True, size=24)
        self.subtitle_font = Font(color=self.accent_color, bold=True, size=14)
        self.bold_font = Font(bold=True)
        
        thin = Side(border_style="thin", color="DDDDDD")
        self.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def generate_excel_quote(self, context: dict) -> str:
        wb = Workbook()
        ws = wb.active
        ws.title = "Devis"

        document_number = context.get("document_number", f"QTE-{str(uuid.uuid4())[:8].upper()}")
        client_name = context.get("client_name", "Client Inconnu")
        items = context.get("items", [])
        
        # 1. Header Section
        ws.merge_cells('A1:E2')
        cell = ws['A1']
        cell.value = "DEVIS COMMERCIAL"
        cell.font = self.title_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
        ws['A4'] = "Numéro de Devis:"
        ws['A4'].font = self.bold_font
        ws['B4'] = document_number
        
        ws['A5'] = "Client:"
        ws['A5'].font = self.bold_font
        ws['B5'] = client_name
        
        # 2. Table Headers
        headers = ["Code", "Description", "Quantité", "Prix Unitaire (MAD)", "Total Ligne (MAD)"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=8, column=col_num)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.border

        # 3. Table Rows
        row_num = 9
        for item in items:
            ws.cell(row=row_num, column=1, value=item.get("code", ""))
            ws.cell(row=row_num, column=2, value=item.get("description", ""))
            ws.cell(row=row_num, column=3, value=float(item.get("quantity", 0)))
            
            price_cell = ws.cell(row=row_num, column=4, value=float(item.get("price", 0)))
            price_cell.number_format = '#,##0.00'
            
            total_cell = ws.cell(row=row_num, column=5, value=float(item.get("line_total", 0)))
            total_cell.number_format = '#,##0.00'
            
            for col in range(1, 6):
                ws.cell(row=row_num, column=col).border = self.border
                if col in [1, 3]:
                    ws.cell(row=row_num, column=col).alignment = Alignment(horizontal="center")
                
            row_num += 1

        # 4. Totals Section
        start_totals = row_num + 2
        totals = [
            ("Sous-total HT:", context.get("original_subtotal", 0.0)),
            ("Remise:", -context.get("discount_amount", 0.0)),
            ("Total HT:", context.get("subtotal", 0.0)),
            ("TVA:", context.get("tax", 0.0)),
            ("Total TTC:", context.get("total", 0.0)),
        ]
        
        for idx, (label, amount) in enumerate(totals):
            current_row = start_totals + idx
            
            label_cell = ws.cell(row=current_row, column=4, value=label)
            label_cell.font = self.bold_font
            label_cell.alignment = Alignment(horizontal="right")
            
            amount_cell = ws.cell(row=current_row, column=5, value=float(amount))
            amount_cell.number_format = '#,##0.00'
            amount_cell.font = self.bold_font
            
            if idx == len(totals) - 1: # Total TTC
                label_cell.fill = self.total_fill
                amount_cell.fill = self.total_fill
                label_cell.font = Font(bold=True, size=12, color=self.primary_color)
                amount_cell.font = Font(bold=True, size=12, color=self.primary_color)

        # 5. Column Width Adjustments
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        
        # 6. Save File
        filename = f"quote_{str(uuid.uuid4())}.xlsx"
        filepath = os.path.join(self.data_dir, filename)
        wb.save(filepath)
        
        return filepath
